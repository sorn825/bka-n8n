import psycopg2
import pandas as pd
import sys
import os

def main():
    # ข้อมูลการเชื่อมต่อ
    conn_params = {
        "host": "postgres",
        "port": "5432",
        "database": "TIVDB",
        "user": "postgres",
        "password": os.environ["POSTGRES_PASSWORD"]
    }

    try:
        # เชื่อมต่อกับ database
        print("กำลังเชื่อมต่อกับฐานข้อมูล...")
        conn = psycopg2.connect(**conn_params)
        
        # Query ข้อมูล
        query = "select * from property_listing order by recorddate desc;"
        print(f"กำลังรัน query: {query}")
        
        # ใช้ cursor อ่านข้อมูลจาก SQL (pandas ไม่รองรับ psycopg2 connection โดยตรง)
        cursor = conn.cursor()
        cursor.execute(query)
        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        df = pd.DataFrame(rows, columns=columns)
        
        # ปิดการเชื่อมต่อ
        cursor.close()
        conn.close()
        
        if df.empty:
            print("ไม่พบข้อมูลในตาราง property_listing")
            return

        # ทำความสะอาดข้อมูล: เปลี่ยน "N/A", "null", "1970-01-01" หรือ NaN/None ให้เป็นค่าว่าง
        import numpy as np
        from datetime import date
        
        # เปลี่ยนค่าที่เป็น string หรือ object ที่ระบุ
        df = df.replace(['N/A', 'null', 'None', '1970-01-01', date(1970, 1, 1)], np.nan)

        # ตรวจสอบเผื่อเป็น datetime object ที่มีค่าเป็น 1970-01-01 (กรณีถูกแปลงเป็น Timestamp)
        for col in df.select_dtypes(include=['datetime64', 'object', 'str']).columns:
            try:
                # ถ้าเป็น datetime column หรือ object ที่อาจเก็บ datetime
                if pd.api.types.is_datetime64_any_dtype(df[col]):
                    df.loc[df[col] == pd.Timestamp('1970-01-01'), col] = np.nan
            except:
                pass

        # --- ส่วนที่เพิ่มใหม่: Merge กับไฟล์เวอร์ชั่นก่อนหน้า ---
        import glob

        # หาไฟล์ property_listing_table_*.xlsx ทั้งหมดในโฟลเดอร์ปัจจุบัน
        try:
            current_dir = "/files/excel"
            os.makedirs(current_dir, exist_ok=True)
            files = sorted(glob.glob(os.path.join(current_dir, "property_listing_table_*.xlsx")))
            
            # ถ้ามีไฟล์เดิมอยู่ (files[-1] คือไฟล์ล่าสุดตาม timestamp)
            if files:
                last_file = files[-1]
                print(f"พบไฟล์เวอร์ชั่นก่อนหน้า: {os.path.basename(last_file)}")
                
                try:
                    df_prev = pd.read_excel(last_file)
                    print(f"โหลดข้อมูลเก่า {len(df_prev)} แถว เพื่อทำการ Merge...")

                    if 'url' in df_prev.columns and 'url' in df.columns:
                        # เก็บชื่อคอลัมน์เดิมเพื่อจัดลำดับและกรองกลับ
                        original_columns = df.columns
                        
                        # ใช้ url เป็น index ในการ merge
                        # หมายเหตุ: เราจะ drop duplicates ใน url ก่อน set index เพื่อป้องกัน error (ถ้ามี)
                        df_new_idx = df.drop_duplicates(subset=['url']).set_index('url')
                        df_prev_idx = df_prev.drop_duplicates(subset=['url']).set_index('url')
                        
                        # combine_first: ใช้ค่าจาก df_new_idx เป็นหลัก ถ้าเป็น NaN ให้ไปเอาจาก df_prev_idx
                        df_merged = df_new_idx.combine_first(df_prev_idx)
                        
                        # กรองเอาเฉพาะ row ที่มีในไฟล์ใหม่ (User Req: ไม่สนใจ row ที่มีในไฟล์เก่าแต่ไม่มีในไฟล์ใหม่)
                        # และกรองเอาเฉพาะ column ที่มีในไฟล์ใหม่ (เพื่อรักษา Structure เดิม)
                        df_final = df_merged.loc[df_new_idx.index, original_columns[original_columns != 'url']]
                        
                        # Reset index เพื่อเอา url กลับมาเป็น column
                        df = df_final.reset_index()
                        
                        # จัดลำดับ column ให้เหมือนเดิม
                        df = df[original_columns]
                        
                        print("Merge ข้อมูลเรียบร้อย (เติมค่าว่างจากข้อมูลเก่า)")
                    else:
                        print("ไม่พบ column 'url' ในไฟล์เก่า หรือไฟล์ใหม่ ข้ามการ Merge")
                        
                except Exception as e:
                    print(f"เกิดข้อผิดพลาดในการอ่านหรือ Merge ไฟล์เก่า: {e}")
            else:
                print("ไม่พบไฟล์เวอร์ชั่นก่อนหน้า เริ่มต้นเป็นไฟล์แรก")
                
        except Exception as e:
            print(f"เกิดข้อผิดพลาดในกระบวนการค้นหาไฟล์เก่า: {e}")
        # -----------------------------------------------------

        # เมื่อใส่ใน excel pandas NaN จะเป็น None ซึ่ง openpyxl จะมองเป็นช่องว่าง
        # เมื่อใส่ใน excel pandas NaN จะเป็น None ซึ่ง openpyxl จะมองเป็นช่องว่าง

        # สร้างชื่อไฟล์ที่มี timestamp
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"/files/excel/property_listing_table_{timestamp}.xlsx"
        
        # ใช้ openpyxl ในการจัดการ table
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.styles import Font, Color
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Property Listing"

        # ใส่ข้อมูลจาก dataframe ลงใน worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # จัดการ Hyperlink สำหรับฟิลด์ที่เป็น URL
        # ข้ามแถวแรกที่เป็น Header
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, str) and (cell.value.startswith('http://') or cell.value.startswith('https://')):
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0563C1", underline="single") # สีน้าเงินมาตรฐาน Hyperlink ใน Excel

        # กำหนดช่วงของ Table
        max_col = get_column_letter(ws.max_column)
        max_row = ws.max_row
        tab_range = f"A1:{max_col}{max_row}"

        # สร้าง Table object
        tab = Table(displayName="PropertyTable", ref=tab_range)

        # เพิ่มสไตล์ให้กับ Table (มี Filter, แถวสลับสี)
        style = TableStyleInfo(
            name="TableStyleMedium9", 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

        # ปรับความกว้างของ Column โดยประมาณ
        for column_cells in ws.columns:
            # คำนวณความยาว ไม่รวม hyperlink metadata
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 60)

        wb.save(filename)
        print(f"ดึงข้อมูลสำเร็จ! บันทึกไฟล์พร้อม Table และ Hyperlink เรียบร้อยที่: {filename}")
        print(f"จำนวนข้อมูลทั้งหมด: {len(df)} แถว")

    except Exception as e:
        print(f"เกิดข้อผิดพลาด: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()

